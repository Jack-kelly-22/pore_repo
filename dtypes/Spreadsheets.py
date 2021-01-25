import os as os
from pandas import DataFrame
from plotly import offline
from openpyxl import Workbook,load_workbook
from math import pi
from PIL import Image
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from plotly.express import bar
import matplotlib.pyplot as plt
from datetime import datetime



class Spreadsheet:
    def __init__(self,frame,job_name):
        self.job_name = job_name
        print("start spread init")
        self.frame_name = frame.name
        self.filename = frame.out_path
        self.workbook = load_workbook("template.xlsx")
        self.i = 1
        self.page = self.workbook.active
        self.col_iter = self.page.iter_cols(min_col=0,max_col=500,max_row=20)
        self.fill_frame_detail(frame)
        self.set_date()
        self.fill_largest_circles(frame)
        self.fill_largest_areas(frame)
        self.fill_histograms(frame)
        self.fill_stds(frame)
        self.fill_overview_by_image(frame)
        self.fill_img_rows(frame)
        self.fill_pie_charts(frame)

        self.workbook.save('.' + frame.out_path + self.job_name  + '/' + frame.name + "_sheet.xlsx")


    
    def fill_largest_areas_header(self):
        self.i = self.i + 1
        #self.style_header(color=, cols=['F', 'G', "H"],sz=12)
        font = Font(name='calibri',
                    size=14,
                    bold=True,
                    color='FFFFFFFF',
                    )
        fill = PatternFill(fill_type='solid', fgColor='00FF8080')
        self.page["F" + str(self.i)].fill = fill
        self.page["F" + str(self.i)].font = font
        self.page["G" + str(self.i)].fill = fill
        self.page["G" + str(self.i)].font = font
        self.page["H" + str(self.i)].fill = fill
        self.page["H" + str(self.i)].font = font
        self.page["F" + str(self.i)] = "Largest Areas in Frame"
        self.i = self.i + 1
        self.page["F" + str(self.i)] = "Image Name"
        self.page["G" + str(self.i)] = "Area(microns)"
        self.page["H" + str(self.i)] = "center(px)"
        self.page["I" + str(self.i)] = "Inscribed Diameter"
        self.i = self.i + 1


    def fill_largest_areas(self,frame):
        self.fill_largest_areas_header()
        sorted_areas = sorted(frame.largest_areas, key=lambda tup: float(tup[1]))
        sorted_areas.reverse()
        i = self.i
        scale = float(frame.constants['scale'])
        # num_rows = frame.constants['num_circles']
        for area in sorted_areas[:10]:
            # output largest holes to sheet
            self.page["F" + str(i)] = area[0]  # Image Name
            self.page["G" + str(i)] = str(round(area[1], 2))  # Area in microns(already scaled)
            self.page["H" + str(i)] = '( ' + str(area[2]) + ' ' + str(area[3]) + ' )'  # center(px)
            #            self.page["I" + str(i)] = str(round(int(int(area[4]) * scale),2))   # diam
            i = i + 1
        if i > self.i:
            self.i = i



    def fill_stds(self,frame):
        font = Font(name='calibri',
                    size=12,
                    bold=True,
                    color='FFFFFFFF',
                    )
        self.page["L5"].font = font
        self.page["N5"].font = font

        self.page["L5"] = str(round(frame.area_std,2))
        self.page["N5"] = str(round(frame.diam_std, 2))

    def set_date(self):
        # datetime object containing current date and time
        now = datetime.now()
        # dd/mm/YY H:M:S
        dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
        self.page["B2"]= dt_string


    def init_cells(self):

        self.page['A' + str(self.i)] = "frame"
        self.page['A' + str(self.i + 1)] = self.frame_name
        self.page['B' + str(self.i)] = "avg porosity"

        self.page["D" + str(self.i)] = "image"
        self.page["E" + str(self.i)] = "Area"
        self.page["F" + str(self.i)] = "Largest circle in area"
        self.page["G" + str(self.i)] = "X"
        self.page["H" + str(self.i)] = "Y"
        self.i = self.i + 1


    def fill_frame_detail(self,frame):
        self.page["C6"] = str(frame.constants["scale"])
        self.page["C7"] = str(len(frame.image_data_ls))
        self.page["C8"] = frame.constants['warn_size']
        self.page["C9"] = frame.constants['min_ignore']
        self.page["C10"] = frame.constants["scale"]
        self.page["C11"] = frame.constants["thresh"]

    def fill_overview_by_image(self,frame):
        i = 17
        print("starting fill overview with i:", self.i )
        for image in frame.image_data_ls:
            self.page["B" + str(i)] = image.name # image Name
            self.page["C" + str(i)] = str(image.porosity) # Porosity
            self.page["D" + str(i)] = str(image.largest_holes[0][1] * 2 * float(image.scale_factor))# Diameter of the largest Circle
            i = i + 1
        if i > self.i:
            self.i = i

    def fill_image_col(self,image_data_ls):
        i = self.i
        for image in image_data_ls:
            self.fill_image_output(image,i)
            i = i + 18


    def fill_image_data_col(self,image_data_ls):
        for image in image_data_ls:
            self.fill_image_table_name(image)
            self.fill_image_table_header(image)
            self.fill_image_table(image)

    def fill_img_rows(self,frame):
        self.fill_image_col(frame.image_data_ls)
        self.fill_image_data_col(frame.image_data_ls)

    def style_header(self,color = '00008080',cols = ['D','E',"F"],sz = 20):

        font = Font(name='calibri',
                    size=sz,
                    bold=True,
                    color='FFFFFFFF',
                    )
        fill_second = PatternFill(fill_type='solid', fgColor=color)
        fill_main = PatternFill(fill_type='solid', fgColor=color)
        for i in range(self.i,self.i+1):
            for row in cols:
                self.page[row + str(i)].font = font
                self.page[row + str(i)].fill = fill_main

    def fill_image_table_name(self,img_data):
        self.style_header(color= '00666699')
        self.page["D" + str(self.i)] = "Image Name: " + img_data.name
        self.i = self.i + 1

    def fill_image_table_header(self,img_data):
        self.style_header()

        self.page["D" + str(self.i)] = "Largest Circles in Image"
        self.i = self.i + 1
        self.page["D" + str(self.i)] = "Diameter(microns)"
        self.page["E" + str(self.i)] = "Area(microns)"
        self.page["F" + str(self.i)] = "center(px)"
        self.i = self.i + 1

    def fill_image_table(self,img_data):
        scale = float(img_data.scale_factor)
        for circle in img_data.largest_holes[:10]:
            # "Diameter(microns)"
            self.page["D" + str(self.i)] = str(circle[1] * scale * 2)
            #Area(microns)
            self.page["E" + str(self.i)] = str(round((circle[1] * circle[1]) * pi * scale * scale, 2))
            #center (px)
            self.page["F" + str(self.i)] =  "(" + str(circle[0][0]) +  "," + str(circle[0][1]) + ")"
            self.i = self.i + 1
        self.i = self.i + 5

    def fill_image_output(self,im_dat,i ):
        og_img = Image('.' +im_dat.image_out_path_og)
        thresh_img = Image('.' + im_dat.image_out_path)
        heat_img = Image('.' + im_dat.heat_out_path)
        og_img.anchor = 'G' + str(i)
        thresh_img.anchor = 'K' + str(i)
        heat_img.anchor = 'P' + str(i)
        og_img.height = 300
        og_img.width = 400
        thresh_img.height = 300
        thresh_img.width = 400
        heat_img.height = 300
        heat_img.width = 400
        self.page.add_image(og_img)
        self.page.add_image(thresh_img)
        self.page.add_image(heat_img)

        #self.i = self.i + 15

    def fill_largest_circles(self,frame):
        sorted_holes = sorted(frame.largest_holes, key=lambda tup:float(tup[1]))
        sorted_holes.reverse()
        i = 7
        scale = float(frame.constants['scale'])
        #num_rows = frame.constants['num_circles']
        for hole in sorted_holes[:10]:
            #output largest holes to sheet
            self.page["F" + str(i)] = hole[2] #Image Name
            self.page["G" + str(i)] = str(round(hole[1] * 2 * scale, 2)) #Diameter in microns
            self.page['H' + str(i)] = str(round((hole[1] * hole[1]) * pi * scale * scale, 2)) #Area microns
            self.page["I" + str(i)] ='( ' + str(hole[0][0]) + ' ' + str(hole[0][1]) + ' )' #center(px)
            i = i + 1
        if i>self.i:
            self.i = i




    def fill_histograms(self,frame):
        hist_paths = [frame.area_hist_path,frame.diam_hist_path]
        col = ['K6','U6']
        j = 0
        for path in hist_paths:
            hist = Image(path)
            hist.anchor = col[j]
            hist.height = 600
            hist.width = 800
            self.page.add_image(hist)
            j = j+1
        if(self.i<34):
            self.i = 34

    def fill_pie_charts(self,frame):
        self.fill_pie_chart_area(frame)
        self.fill_pie_chart_pores(frame)

    def fill_pie_chart_area(self,frame):
        pie = Image(frame.area_pie_path)
        pie.anchor = "AE35"
        pie.height = 600
        pie.width = 800
        self.page.add_image(pie)

    def fill_pie_chart_pores(self,frame):
        pie = Image(frame.pore_pie_path)
        pie.anchor = "AE6"
        pie.height = 600
        pie.width = 800
        self.page.add_image(pie)


